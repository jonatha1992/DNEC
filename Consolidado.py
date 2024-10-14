import openpyxl

# Carga el archivo de Excel
wb_base = openpyxl.load_workbook('data/Base de datos.xlsx')
wb_procedimientos= openpyxl.load_workbook('bajadas/bajada_general.xls')
wb_arma= openpyxl.load_workbook('bajadas/bajada_arma.xls')
wb_vehiculos_secuestros= openpyxl.load_workbook('bajadas/bajada_vehiculos_secuestrados.xls')
wb_vehiculos_secuestros_ministerio= openpyxl.load_workbook('bajadas/bajada_ministerio_vehiculos.xls')
wb_personas= openpyxl.load_workbook('bajadas/bajada_general_persona.xls')
wb_secuestradas= openpyxl.load_workbook('bajadas/bajada_divisas_secuestradas.xls')
wb_sustancia= openpyxl.load_workbook('bajadas/bajada_general_narcotrafico.xls')

# Selecciona la primera hoja de cálculo
libro_base = wb_base['Base']
libro_procedimientos = wb_procedimientos['Page 1']
libro_armas = wb_arma['Report']
libro_vehiculos_1 = wb_vehiculos_secuestros['Report']
libro_vehiculos_2 = wb_vehiculos_secuestros_ministerio['Report']
libro_persona = wb_personas['Report']
libro_divisas = wb_secuestradas['Report']
libro_sustancias = wb_sustancia['Report']



# Cierra el archivo de Excel