import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from PyQt6.QtCore import QObject, pyqtSignal
from funciones import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
pd.set_option('future.no_silent_downcasting', True)

PATH_BASE = 'db/base_informada.xlsx'
PATH_TEMPLATE = 'models/modelo_informe.xlsx'
PATH_FILE_OUTPUT = ''
# Configuración para evitar el FutureWarning

class Controlador(QObject):
    # Señales para comunicar progreso
    progress = pyqtSignal(int)  # For progress value and status message
    status = pyqtSignal(str)
    finished = pyqtSignal(bool)
    error = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.archivos = {}
        self.fecha_inicial = None
        self.fecha_final = None
        self.verificar_base = False
        self.CONTADOR = self.inicializar_contador()
        
    def inicializar_contador(self):
        """Inicializa el diccionario contador"""
        return {
            "FECHA_MENOR_BASE":datetime,
            "FECHA_MAYOR_BASE":datetime,
            'PROCEDIMIENTOS_NUEVOS': 0,
            'ORDEN_SERVICIOS_NUEVOS': 0,
            'BAJADA_ORDEN_SERVICIOS': 0,
            'BAJADA_PROCEDIMIENTOS': 0,
            'BAJADA_ARMAS': 0,
            'BAJADA_DIVISAS': 0,
            'BAJADA_NARCOTRAFICO': 0,
            'BAJADA_OBJETOS': 0,
            'BAJADA_PERSONAS': 0,
            'BAJADA_VEHICULOS': 0,
            'BAJADA_TRATA': 0,
            'GEOG_FINAL': 0,
            'VICTIMAS_FINAL': 0,
            'DETENIDOS_FINAL': 0,
            'ARMAS_FINAL': 0,
            'DIVISAS_FINAL': 0,
            'NARCOTRAFICO_FINAL': 0,
            'OBJETOS_FINAL': 0,
            'VEHICULOS_FINAL': 0,
            'TRATA_FINAL': 0,
        }

    def validar_fechas(self, fecha_inicial, fecha_final):
        """Valida que las fechas sean correctas"""
        try:
            if fecha_inicial > fecha_final:
                raise ValueError("La fecha inicial no puede ser mayor que la fecha final")
            return True
        except Exception as e:
            raise Exception(f"Error en las fechas: {str(e)}")
            
    def validar_archivos(self, archivos):
        """Valida que los archivos necesarios estén presentes"""
        tipos_requeridos = ['procedimiento', 'arma', 'divisa', 'narcotrafico', 
                          'objetos', 'persona', 'vehiculo', 'oper']
        
        archivos_encontrados = {tipo: False for tipo in tipos_requeridos}
        
        for archivo in archivos:
            nombre = archivo.lower()
            for tipo in tipos_requeridos:
                if tipo in nombre:
                    archivos_encontrados[tipo] = True
                    
        faltantes = [tipo for tipo, encontrado in archivos_encontrados.items() 
                    if not encontrado]
        
        if faltantes:
            raise Exception(f"Faltan archivos requeridos: {', '.join(faltantes)}")
        
        return True

    def iniciar_procesamiento(self, archivos, fecha_inicial, fecha_final, verificar_base=False):
        """Inicia el proceso de consolidación"""
        try:
            # Validaciones iniciales
            self.validar_fechas(fecha_inicial, fecha_final)
            self.validar_archivos(archivos)
            
            # Guardar parámetros
            self.archivos = self.clasificar_archivos(archivos)
            self.fecha_inicial = fecha_inicial.strftime('%d-%m-%Y')
            self.fecha_final = fecha_final.strftime('%d-%m-%Y')
            self.verificar_base = verificar_base

            self.progress.emit(10)
            self.status.emit("Procesando procedimientos...")
            df_procedimientos = self.procesar_procedimientos()
            
            self.progress.emit(20)
            if self.CONTADOR['PROCEDIMIENTOS_NUEVOS'] > 0:
                self.status.emit("Procesando personas...")
                df_detenidos , df_otros_delitos = self.procesar_personas(df_procedimientos)
                
                self.progress.emit(30)
                self.status.emit("Procesando incautaciones...")
                df_incautaciones = self.procesar_incautaciones(df_procedimientos)
                
                self.progress.emit(50)
                self.status.emit("Procesando trata...")
                df_trata = self.procesar_trata(df_procedimientos)
            else:
                self.progress.emit(50)
                self.status.emit("No hay procedimientos nuevos para procesar")
                df_detenidos = pd.DataFrame()
                df_otros_delitos = pd.DataFrame()
                df_incautaciones = pd.DataFrame()
                df_trata = pd.DataFrame()
            
            
            
            self.progress.emit(60)
            self.status.emit("Procesando operaciones...")
            df_operaciones, df_controlados, df_afectados, df_codigos = self.procesar_operaciones()
    
            self.progress.emit(70)
            self.status.emit("Consolidando datos...")
            dataframes = self.consolidar_datos( df_procedimientos, df_operaciones, df_incautaciones, df_detenidos, df_otros_delitos, df_trata , df_afectados, df_controlados, df_codigos)
            
            self.progress.emit(80)
            self.status.emit("Ordenando...")
            dataframes = self.ordenar_columnas( dataframes)
            
            self.progress.emit(90)
            self.status.emit("Generando informe...")
            self.copiar_formato_template(dataframes)
            
            self.progress.emit(100)
            self.status.emit("Procesamiento completado")
            
            return True
            
        except Exception as e:
            self.error.emit(f"Error en el procesamiento: {str(e)}")
            raise
        
        
    def filtrar_con_base(self, df : pd.DataFrame) -> pd.DataFrame:
        """
        Filters dataframes against the base database using ID_PROCEDIMIENTO 
        to remove already reported records. Tracks counts before and after filtering.
        """
        # Read existing records from base
        df_base = pd.read_excel(PATH_BASE, sheet_name="GEOG. PROCEDIMIENTO")
        
        # Create set of existing IDs for faster lookup
        ids_base = set(df_base['ID_PROCEDIMIENTO'])
        
        # Track counts for reporting
        if len(ids_base) > 0:
            # Filter out existing IDs
            df = df[~df['ID_PROCEDIMIENTO'].isin(ids_base)]
            
        return df
    

    def clasificar_archivos(self, archivos):
        """Clasifica los archivos según su tipo"""
        clasificados = {
            'procedimiento': None,
            'persona':None,
            'arma': None,
            'divisas': None,
            'narcotrafico': None,
            'objetos': None,
            'vehiculos': None,
            'operaciones': None,
            'trata': None
        }
        
        for archivo in archivos:
            nombre = archivo.lower()
            for tipo in clasificados.keys():
                if tipo in nombre:
                    clasificados[tipo] = archivo
                    break
                    
        return clasificados

    def procesar_procedimientos(self):
        """Procesa los archivos de procedimientos"""
        
        # df = pd.read_excel(self.archivos['procedimiento'])
        df = filtrar_procedimientos_generales(self.archivos['procedimiento'])
        self.progress.emit(15)
        self.status.emit( "Procesando procedimientos...")
        df_procedimientos = pd.DataFrame()
        df_procedimientos['ID_OPERATIVO'] = df.apply(procesar_causa_judicial, axis=1)
        df_procedimientos['FUERZA_INTERVINIENTE'] = "PSA"
        df_procedimientos['ID_PROCEDIMIENTO'] = df.apply(generar_uid_sigpol, axis=1)
        df_procedimientos['CAUSAJUDICIALNUMERO'] = df['CAUSAJUDICIALNUMERO'].copy()
        df_procedimientos['UNIDAD_INTERVINIENTE'] = df['UOSP'].fillna(df['URSA'])
        df_procedimientos['DESCRIPCIÓN'] = df.apply(procesar_descripcion, axis=1)
        df_procedimientos['TIPO_INTERVENCION'] = df.apply(procesar_tipo_procedimiento, axis=1)
        df_procedimientos['PROVINCIA'] = df.apply(procesar_provincia, axis=1)
        df_procedimientos['DEPARTAMENTO O PARTIDO'] = df.apply(procesar_municipio, axis=1)
        df_procedimientos['LOCALIDAD'] = "-"
        df_procedimientos['DIRECCION'] = df.apply(procesar_direccion, axis=1)
        df_procedimientos[['LATITUD', 'LONGITUD']] = df.apply(procesar_geog, axis=1, result_type='expand')
        df_procedimientos['FECHA'] = pd.to_datetime(df['DENUNCIAFECHA'], errors='coerce').dt.date
        df_procedimientos['HORA'] = pd.to_datetime(df['DENUNCIAFECHA'], errors='coerce').dt.strftime('%H:%M')
        df_procedimientos['ZONA_SEGURIDAD_FRONTERAS'] = "-"
        df_procedimientos['PASO_FRONTERIZO'] = "-"
        df_procedimientos['OTRAS AGENCIAS INTERVINIENTES'] = "-"
        df_procedimientos['Observaciones - Detalles'] = "-"
        df_procedimientos['LATITUD'] = df_procedimientos['LATITUD'].astype(str).str.replace(',', '.')
        df_procedimientos['LONGITUD'] = df_procedimientos['LONGITUD'].astype(str).str.replace(',', '.')

        
        if self.verificar_base == True:
            df_procedimientos = self.filtrar_con_base(df_procedimientos)
        
        df_procedimientos_completado = df_procedimientos[['FUERZA_INTERVINIENTE', 'ID_OPERATIVO', 'ID_PROCEDIMIENTO',
                                            'UNIDAD_INTERVINIENTE', 'DESCRIPCIÓN', 'TIPO_INTERVENCION',
                                            'PROVINCIA', 'DEPARTAMENTO O PARTIDO', 'LOCALIDAD', 'DIRECCION',
                                            'ZONA_SEGURIDAD_FRONTERAS', 'PASO_FRONTERIZO', 'LATITUD', 'LONGITUD',
                                            'FECHA', 'HORA', 'OTRAS AGENCIAS INTERVINIENTES', 'Observaciones - Detalles']]
        
        
        self.CONTADOR['BAJADA_PROCEDIMIENTOS'] = len(df)
        self.CONTADOR['PROCEDIMIENTOS_NUEVOS'] = len(df_procedimientos)
        
        return df_procedimientos_completado

    def procesar_personas(self, df_procedimientos):
        """Procesa los archivos de personas"""
        df = pd.read_excel(self.archivos['persona'])
        
        if df.empty:
            return pd.DataFrame()
            
        self.CONTADOR['BAJADA_PERSONAS'] = len(df)
        
        # Generar UID y procesar datos
        df['UOSP'] = df['UOSP'].fillna(df['URSA'])
        df['ID_PROCEDIMIENTO'] = df.apply(generar_uid_sigpol, axis=1)
        # Paso 3: Coincidencia de UID entre excel_bajada_personas y df_geog_final
        uid_coincidentes = df[df['ID_PROCEDIMIENTO'].isin(df_procedimientos['ID_PROCEDIMIENTO'])]

        # Crear DataFrame de detenidos y aprendidos
        df_detenidos_aprendidos = pd.DataFrame()
        df_detenidos_aprendidos['ID_PROCEDIMIENTO'] = uid_coincidentes['ID_PROCEDIMIENTO']
        df_detenidos_aprendidos['EDAD'] = uid_coincidentes.apply(procesar_edad, axis=1)
        df_detenidos_aprendidos['SEXO'] = uid_coincidentes.apply(procesar_sexo, axis=1)
        df_detenidos_aprendidos['GENERO'] = uid_coincidentes.apply(procesar_genero, axis=1)
        df_detenidos_aprendidos['NACIONALIDAD'] = uid_coincidentes.apply(procesar_nacionalidad, axis=1)
        df_detenidos_aprendidos['SITUACION_PROCESAL'] = uid_coincidentes.apply(procesar_situacion_judicial, axis=1)
        df_detenidos_aprendidos['DELITO_IMPUTADO'] = uid_coincidentes.apply(procesar_tipo_delito, axis=1)
        df_detenidos_aprendidos['JUZGADO_INTERVINIENTE'] = uid_coincidentes.apply(procesar_juzgado, axis=1)
        df_detenidos_aprendidos['CARATULA_CAUSA'] = uid_coincidentes.apply(procesar_caratula, axis=1)
        df_detenidos_aprendidos['NUM_CAUSA'] = uid_coincidentes['ID_PROCEDIMIENTO']


        df_detenidos_aprendidos = df_detenidos_aprendidos[df_detenidos_aprendidos['SITUACION_PROCESAL'] != 'NO INFORMAR']
        df_detenidos_aprendidos = df_detenidos_aprendidos[df_detenidos_aprendidos['ID_PROCEDIMIENTO'].isin(df_procedimientos['ID_PROCEDIMIENTO'])]
        print(df_detenidos_aprendidos.nunique())
        
        
    
        if not df.empty:
            
            df_otros_delitos = df_detenidos_aprendidos[df_detenidos_aprendidos['SITUACION_PROCESAL'] == 'VICTIMA']
            # Filtro los que son victima de detenidos y aprendidos
            df_detenidos_aprendidos_completado = df_detenidos_aprendidos[df_detenidos_aprendidos['SITUACION_PROCESAL'] != 'VICTIMA']

            # Cambiar el nombre de algunas columnas en el DataFrame de víctimas
            df_otros_delitos_completado = df_otros_delitos.rename(columns={
                'DELITO_IMPUTADO': 'TIPO_OTRO_DELITO',
                'GENERO': 'GENERO_VICTIMA',
                'EDAD': 'EDAD_VICTIMA',
            }) # type: ignore
            df_otros_delitos_completado["OBSERVACIONES"] ="-"
            
            self.CONTADOR['DETENIDOS_FINAL'] = len(df_detenidos_aprendidos_completado['ID_PROCEDIMIENTO'])
            self.CONTADOR['VICTIMAS_FINAL'] = len(df_otros_delitos_completado['ID_PROCEDIMIENTO'])
            
            print(df_otros_delitos_completado.nunique())
                
        return df_detenidos_aprendidos_completado, df_otros_delitos_completado

    def procesar_incautaciones(self, df_procedimientos):
        
        """Procesa todas las incautaciones"""
        self.progress.emit(52)
        self.status.emit( "Procesando armas...")
        df_armas = self.procesar_armas(df_procedimientos)
        
        self.progress.emit(54)
        self.status.emit( "Procesando divisas...")
        df_divisas = self.procesar_divisas(df_procedimientos)
        
        self.progress.emit(56)
        self.status.emit( "Procesando objetos...")
        df_objetos = self.procesar_objetos(df_procedimientos)
        
        self.progress.emit(58)
        self.status.emit( "Procesando vehículos...")
        df_vehiculos = self.procesar_vehiculos(df_procedimientos)
        
        self.progress.emit(60)
        self.status.emit( "Procesando narcotráfico...")
        df_narcotrafico = self.procesar_narcotrafico(df_procedimientos)
        
        df_incautaaciones = pd.concat([df_armas, df_divisas, df_objetos, df_vehiculos, df_narcotrafico])

        df_incautaaciones = df_incautaaciones[df_incautaaciones['ID_PROCEDIMIENTO'].isin(df_procedimientos['ID_PROCEDIMIENTO'])]
        return df_incautaaciones 

    def procesar_armas(self, df_procedimientos):
        """Procesa incautaciones de armas"""
        df = pd.read_excel(self.archivos['arma'])
        if df.empty:
            return pd.DataFrame()
            
        self.CONTADOR['BAJADA_ARMAS'] = len(df)

        df = df[df["TIPO_ESTADO_OBJETO"] == "SECUESTRADO"]
        df['UOSP'] = df['UOSP'].fillna(df['URSA'])
        df['TIPO_CAUSA_INTERNA'] = df.apply( procesar_tipo_causa_interna, axis=1)
        df['ID_PROCEDIMIENTO'] = df.apply(generar_uid_sigpol, axis=1)


        df_arma = pd.DataFrame()
        df_arma['ID_PROCEDIMIENTO'] = df['ID_PROCEDIMIENTO']
        df_arma['TIPO_INCAUTACION'] = "ARMAS"
        df_arma['TIPO'] = df['TIPO_ARMA']
        df_arma['SUBTIPO'] = "-"
        df_arma['CANTIDAD'] = df.apply(procesar_cantidad_arma, axis=1)
        df_arma['MEDIDAS'] = "UNIDADES"
        df_arma['AFORO'] = "-"
        df_arma['OBSERVACIONES'] = df.apply(procesar_observaciones_arma, axis=1)

        df_arma_completado = df_arma[df_arma['ID_PROCEDIMIENTO'].isin(df_procedimientos['ID_PROCEDIMIENTO'])]

        self.CONTADOR['ARMAS_FINAL'] = len(df_arma['ID_PROCEDIMIENTO'])
        print(df_arma_completado.nunique())

        return df_arma_completado

    def procesar_divisas(self, df_procedimientos):
        """Procesa incautaciones de divisas"""
        df = pd.read_excel(self.archivos['divisas'])
        excel_bajada_procedimientos_generales = pd.read_excel(self.archivos['procedimiento'])
        if df.empty:
            return pd.DataFrame()
            
        self.CONTADOR['BAJADA_DIVISAS'] = len(df)
        

        excel_bajada_procedimientos_generales['PARTE_ANIO'] = excel_bajada_procedimientos_generales['NUMERO_PARTE'].astype(str) + "/" + excel_bajada_procedimientos_generales['ANIO_PARTE'].astype(str)
        df['PARTE_ANIO'] = df['NUMERO_PARTE'].astype(str) + "/" + df['ANIO_PARTE'].astype(str)

        df['UOSP'] = df['UOSP'].fillna(df['URSA'])

        df = pd.merge(df, 
                                    excel_bajada_procedimientos_generales[['PARTE_ANIO', 'TIPO_CAUSA_INTERNA']], 
                                    on='PARTE_ANIO', 
                                    how='left')

        df['ID_PROCEDIMIENTO'] = df.apply(generar_uid_sigpol, axis=1)
        df = df[~pd.isnull(df['TOTAL_DIVISAS_SECUESTRADAS'])]


        cantidad_por_uid = df.groupby(['ID_PROCEDIMIENTO', 'TIPO_DIVISA'], as_index=False)['TOTAL_DIVISAS_SECUESTRADAS'].sum()
        

        df_divisa = pd.DataFrame()
        df_divisa['ID_PROCEDIMIENTO'] = df['ID_PROCEDIMIENTO']
        df_divisa['TIPO_INCAUTACION'] = "DIVISAS"
        df_divisa['TIPO'] = cantidad_por_uid['TIPO_DIVISA']
        df_divisa['SUBTIPO'] = "-"
        df_divisa['CANTIDAD'] = cantidad_por_uid['TOTAL_DIVISAS_SECUESTRADAS']
        df_divisa['MEDIDAS'] = "UNIDADES"
        df_divisa['AFORO'] ="-"
        df_divisa['OBSERVACIONES'] = "-"

        df_divisa_completado = df_divisa[df_divisa['ID_PROCEDIMIENTO'].isin(df_procedimientos['ID_PROCEDIMIENTO'])]

        self.CONTADOR['DIVISAS_FINAL'] = len(df_divisa_completado['ID_PROCEDIMIENTO'])
        
        print(df_divisa_completado.nunique())

        return df_divisa_completado
        

    def procesar_objetos(self, df_procedimientos):
        """Procesa incautaciones de objetos"""
        df = pd.read_excel(self.archivos['objetos'])
        if df.empty:
            return pd.DataFrame()
            
        self.CONTADOR['BAJADA_OBJETOS'] = len(df)


        df = df[df["ESTADO"] == "SECUESTRADO"]
        df['TIPO_OBJETO'] = df['TIPO_OBJETO'].str.replace('Ã“', 'Ó')
        df['UOSP'] = df['UOSP'].fillna(df['URSA'])
        df['TIPO_CAUSA_INTERNA'] = df.apply(procesar_tipo_causa_interna, axis=1)
        df['ID_PROCEDIMIENTO'] = df.apply(generar_uid_sigpol, axis=1)
        df = df[df['TIPO_OBJETO'].isin(TIPO_OBJETO)]
        df = df[~pd.isnull(df['CANTIDAD'])]

        df_objetos = pd.DataFrame()
        df_objetos['ID_PROCEDIMIENTO'] = df['ID_PROCEDIMIENTO']
        df_objetos['TIPO_INCAUTACION'] = "MERCADERIA"
        df_objetos['TIPO'] = df['TIPO_OBJETO'] 
        df_objetos['SUBTIPO'] = "-"
        df_objetos['CANTIDAD'] = df['CANTIDAD']
        df_objetos['MEDIDAS'] = "UNIDADES"
        df_objetos['AFORO'] = "-"
        df_objetos['OBSERVACIONES'] = df['OBJETO']


        df_objetos_completado = df_objetos[df_objetos['ID_PROCEDIMIENTO'].isin(df_procedimientos['ID_PROCEDIMIENTO'])]

        self.CONTADOR['OBJETOS_FINAL'] = len(df_objetos_completado['ID_PROCEDIMIENTO'])
        print(df_objetos_completado.nunique())

        return df_objetos_completado

    def procesar_vehiculos(self, df_procedimientos):
        #* """Procesa incautaciones de vehículos"""
        df = pd.read_excel(self.archivos['vehiculos'])
        if df.empty:
            return pd.DataFrame()
            
        self.CONTADOR['BAJADA_VEHICULOS'] = len(df)

        df = df[df["VEHICULO_ESTADO"] == "SECUESTRADO"]
        df['UOSP'] = df['UOSP'].fillna(df['URSA'])
        df['TIPO_CAUSA_INTERNA'] = df.apply(procesar_tipo_causa_interna, axis=1)
        df['ID_PROCEDIMIENTO'] = df.apply(generar_uid_sigpol, axis=1)

        df_vehiculos = pd.DataFrame()
        df_vehiculos['ID_PROCEDIMIENTO'] = df['ID_PROCEDIMIENTO']
        df_vehiculos['TIPO_INCAUTACION'] = "SECUESTRO DE VEHICULOS"
        df_vehiculos['TIPO'] = df['VEHICULO_TIPO']
        df_vehiculos['SUBTIPO'] = "-"
        df_vehiculos['CANTIDAD'] = 1
        df_vehiculos['MEDIDAS'] = "UNIDADES"
        df_vehiculos['AFORO'] = "-"
        df_vehiculos['OBSERVACIONES'] = df.apply(observaciones_vehiculo, axis=1)

        df_vehiculos_completado = df_vehiculos[df_vehiculos['ID_PROCEDIMIENTO'].isin(df_procedimientos['ID_PROCEDIMIENTO'])]

        self.CONTADOR['VEHICULOS_FINAL'] = len(df_vehiculos_completado['ID_PROCEDIMIENTO'])

        return df_vehiculos_completado

    def procesar_narcotrafico(self, df_procedimientos):
        """Procesa incautaciones de narcotráfico"""
        df = pd.read_excel(self.archivos['narcotrafico'])
        if df.empty:
            return pd.DataFrame()
            
        self.CONTADOR['BAJADA_NARCOTRAFICO'] = len(df)

        df['UOSP'] = df['UOSP'].fillna(df['URSA'])
        df['TIPO_CAUSA_INTERNA'] = df.apply(procesar_tipo_causa_interna, axis=1)
        df['ID_PROCEDIMIENTO'] = df.apply(generar_uid_sigpol, axis=1)
        df = df[~pd.isnull(df['TIPO_ESTUPEFACIENTE'])]

        df_narcotrafico = pd.DataFrame()
        df_narcotrafico['ID_PROCEDIMIENTO'] = df['ID_PROCEDIMIENTO']
        df_narcotrafico['TIPO_INCAUTACION'] = "ESTUPEFACIENTE"
        df_narcotrafico['TIPO'] = df.apply(clasificar_tipo_sustancia, axis=1)
        df_narcotrafico['SUBTIPO'] = "-"
        df_narcotrafico[['CANTIDAD', 'MEDIDAS']] = df.apply(clasificar_medida, axis=1, result_type='expand')
        df_narcotrafico['AFORO'] = "-"
        df_narcotrafico['OBSERVACIONES'] = df.apply(observaciones_sustancia, axis=1)

        df_narcotrafico_completado = df_narcotrafico[df_narcotrafico['ID_PROCEDIMIENTO'].isin(df_procedimientos['ID_PROCEDIMIENTO'])]
        self.CONTADOR['NARCOTRAFICO_FINAL'] = len(df_narcotrafico_completado['ID_PROCEDIMIENTO'])

        return df_narcotrafico_completado

    def procesar_operaciones(self):
        """Procesa las operaciones y órdenes de servicio"""
        
        self.progress.emit(72)
        self.status.emit( "Procesando operaciones...")
        df = pd.read_excel(self.archivos['operaciones'], sheet_name="ORDEN_SERVICIOS", skiprows=1)
        self.CONTADOR['BAJADA_ORDEN_SERVICIOS'] = len(df)
        
        
        if self.verificar_base == True:
            df = self.filtrar_con_base(df)
        
        self.CONTADOR['ORDEN_SERVICIOS_NUEVOS'] = len(df)
        if self.CONTADOR['ORDEN_SERVICIOS_NUEVOS'] == 0:
            
            self.error.emit("No hay órdenes de servicio nuevas para procesar") 
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
        
        df_operaciones = self.procesar_ordenes_servicios(df)
        df_controlados = self.procesar_controlados(df)
        df_afectados = self.procesar_afectados(df)
        df_codigos = self.procesar_codigos_operativos(df)
        
        return df_operaciones, df_controlados, df_afectados, df_codigos


    def procesar_ordenes_servicios(self, df):
        """Procesa los datos de vehículos y personas controladas"""
        
        # Usar .strftime('%H:%M') en cada valor para obtener solo la hora y el minuto
        if df.empty:
            return pd.DataFrame()
        df['HORA'] = df['HORA'].apply(lambda x: x.strftime('%H:%M') if pd.notnull(x) else None)
        df["PROVINCIA"] = df["PROVINCIA"].astype(str).str.strip()
        
        
        df_operaciones= pd.DataFrame()
        df_operaciones["ID_PROCEDIMIENTO"] = df["ID_PROCEDIMIENTO"]
        df_operaciones["FUERZA_INTERVINIENTE"] = "PSA"
        df_operaciones["ID_OPERATIVO"] = df["ID_OPERATIVO"]
        df_operaciones["UNIDAD_INTERVINIENTE"] = df["UNIDAD_INTERVINIENTE"]
        df_operaciones["DESCRIPCIÓN"] = df["DESCRIPCIÓN"]
        df_operaciones["TIPO_INTERVENCION"] = df["TIPO_INTERVENCION"]
        df_operaciones["PROVINCIA"] = df["PROVINCIA"].str.replace("_", " ", regex=False)
        df_operaciones["DEPARTAMENTO O PARTIDO"] = df["DEPARTAMENTO O PARTIDO"].str.upper()
        df_operaciones["LOCALIDAD"] = df["LOCALIDAD"]
        df_operaciones["DIRECCION"] = df["DIRECCION"]
        df_operaciones['FECHA'] = pd.to_datetime(df["FECHA"]).dt.date
        df_operaciones['HORA'] = df['HORA']
        df_operaciones["ZONA_SEGURIDAD_FRONTERAS"] = "-"
        df_operaciones["PASO_FRONTERIZO"] = "-"
        df_operaciones['OTRAS AGENCIAS INTERVINIENTES'] =  df["OTRAS AGENCIAS INTERVINIENTES"]
        df_operaciones['Observaciones - Detalles'] = "-"
        df_operaciones[['LATITUD', 'LONGITUD']] = df.apply(procesar_geog_oper, axis=1, result_type='expand')
            
        
        # Reemplazar "S/D" y "N/C" por "-"
        df_operaciones.replace(["S/D", "N/C"], "-", inplace=True)

        # Reemplazar los valores vacíos (NaN) por "-"
        df_operaciones.fillna("-", inplace=True)

        df_operaciones = df_operaciones[df_operaciones['ID_PROCEDIMIENTO'] != "-"]

        return df_operaciones
    
    def procesar_controlados(self, df):
        """Procesa los datos de vehículos y personas controladas"""
        if df.empty:
            return pd.DataFrame()
            
        df_controlados = pd.DataFrame()
        df_controlados["ID_PROCEDIMIENTO"] = df["ID_PROCEDIMIENTO"]
        df_controlados["FUERZA_INTERVINIENTE"] = "PSA"
        df_controlados["ID_OPERATIVO"] = df["ID_OPERATIVO"]
        df_controlados["UNIDAD_INTERVINIENTE"] = df["UNIDAD_INTERVINIENTE"]
        df_controlados["DESCRIPCIÓN"] = df["DESCRIPCIÓN"]
        df_controlados["TIPO_INTERVENCION"] = df["DESCRIPCIÓN"]
        df_controlados["VEHICULOS_CONTROLADOS"] = df["VEHICULOS_CONTROLADOS"]
        df_controlados["PERSONAS_CONTROLADAS"] = df["PERSONAS_CONTROLADAS"]
        df_controlados["CANT_AVERIGUACIONES_SECUESTRO"] = df["CANT_AVERIGUACIONES_SECUESTRO"]
        df_controlados["CANT_SOLICITUDES_ANTECEDENTES"] = df["CANT_SOLICITUDES_ANTECEDENTES"]
        
        
        # Reemplazar "S/D" y "N/C" por "-"
        df_controlados.replace(["S/D", "N/C"], "-", inplace=True)
        
        return df_controlados

    def procesar_afectados(self, df):
        """Procesa los datos del personal y elementos afectados"""
        if df.empty:
            return pd.DataFrame()
            
        df_afectados = pd.DataFrame()
        df_afectados["FUERZA_INTERVINIENTE"] = "PSA"
        df_afectados["ID_PROCEDIMIENTO"] = df["ID_PROCEDIMIENTO"]
        df_afectados["ID_OPERATIVO"] = df["ID_OPERATIVO"]
        df_afectados["UNIDAD_INTERVINIENTE"] = df["UNIDAD_INTERVINIENTE"]
        df_afectados["DESCRIPCIÓN"] = df["DESCRIPCIÓN"]
        df_afectados["TIPO_INTERVENCION"] = df["TIPO_INTERVENCION"]
        df_afectados["CANT_EFECTIVOS"] = df["CANT_EFECTIVOS"]
        df_afectados["CANT_AUTOS_CAMIONETAS"] = df["CANT_AUTOS_CAMIONETAS"]
        df_afectados["CANT_SCANNERS"] = df["CANT_SCANNERS"]
        df_afectados["CANT_EMBARCACIONES"] = df["CANT_EMBARCACIONES"]
        df_afectados["CANT_MOTOS"] = df["CANT_MOTOS"]
        df_afectados["CANT_CABALLOS"] = df["CANT_CABALLOS"]
        df_afectados["CANT_CANES"] = df["CANT_CANES"]
        df_afectados["CANT_MORPHRAPID"] = df["CANT_MORPHRAPID"]
        df_afectados["CANT_LPR"] = df["CANT_LPR"]
        
        # Reemplazar "S/D" y "N/C" por "-"
        df_afectados.replace(["S/D", "N/C"], "-", inplace=True)

        # Reemplazar los valores vacíos (NaN) por "-"
        df_afectados.fillna("-", inplace=True)

        return df_afectados



    def procesar_codigos_operativos(self, df):
        """Procesa los códigos operativos"""
        if df.empty:
            return pd.DataFrame()
            
        df_codigo = pd.DataFrame()
        df_codigo["UID"] = df["ID_PROCEDIMIENTO"]
        df_codigo["FUERZA_INTERVINIENTE"] = "PSA"
        df_codigo["ID_PROCEDIMIENTO"] = df["ID_PROCEDIMIENTO"]
        df_codigo["ID_OPERATIVO"] = df["ID_OPERATIVO"]
        df_codigo["UNIDAD_INTERVINIENTE"] = df["UNIDAD_INTERVINIENTE"]
        df_codigo["DESCRIPCIÓN"] = df["DESCRIPCIÓN"]
        df_codigo["TIPO_INTERVENCION"] = df["TIPO_INTERVENCION"]
        df_codigo["CODIGO_OPERATIVO"] = df["CODIGO_OPERATIVO"]

        # Reemplazar "S/D" y "N/C" por "-"
        df_codigo.replace(["S/D", "N/C"], "-", inplace=True)

        # Reemplazar los valores vacíos (NaN) por "-"
        df_codigo.fillna("-", inplace=True)

        return df_codigo

    def procesar_trata(self, df_procedimientos):
        """Procesa casos de trata de personas"""
        self.progress.emit(82)
        self.progress.emit( "Procesando casos de trata...")
        df = pd.read_excel(self.archivos['trata'])
        
        if df.empty:
            return pd.DataFrame()
            
        self.CONTADOR['BAJADA_TRATA'] = len(df)

        df['UOSP'] = df['UOSP'].fillna(df['URSA'])
        df['TIPO_CAUSA_INTERNA'] = df.apply( procesar_tipo_causa_interna, axis=1)
        df['ID_PROCEDIMIENTO'] = df.apply(generar_uid_sigpol, axis=1)
        
        df_trata = pd.DataFrame()
        df_trata['ID_PROCEDIMIENTO'] = df['ID_PROCEDIMIENTO']
        df_trata['TIPO_DELITO'] = df["TIPO_EXPLOTACION"].fillna("-")
        df_trata['SEXO_VICTIMA'] = df.apply(procesar_sexo, axis=1)
        df_trata['GENERO_VICTIMA'] = df.apply(procesar_genero, axis=1)
        df_trata['EDAD_VICTIMA'] = df['EDAD'].fillna("-")
        df_trata['NACIONALIDAD'] = df['NACIONALIDAD'].fillna("-")
        df_trata['NACIONALIDAD'] = df['NACIONALIDAD'].str.upper()
        df_trata['JUZGADO_INTERVINIENTE'] = df.apply(procesar_juzgado, axis=1)
        df_trata['OBSERVACIONES'] = "-"
        
        df_trata_completado = df_trata[df_trata['ID_PROCEDIMIENTO'].isin(df_procedimientos['ID_PROCEDIMIENTO'])]
        df_trata_completado = df_trata[df_trata['TIPO_DELITO'] != "-"]
        
        self.CONTADOR['TRATA_FINAL'] = len(df_trata_completado['ID_PROCEDIMIENTO'])
        
        return df_trata_completado
        
    def consolidar_datos(self, df_procedimientos, df_operaciones, df_incautaciones, df_detenidos, df_otros_delitos, df_trata, df_afectados, df_controlados, df_codigos):
        """Consolida todos los datos procesados"""
        # Unir datos de procedimientos y operaciones
        df_geog_final = pd.concat([df_procedimientos, df_operaciones])
        self.CONTADOR["GEOG_FINAL"] = len(df_geog_final)

        # Realizar los merge necesarios
        if self.CONTADOR['PROCEDIMIENTOS_NUEVOS'] > 0:
            df_incautados_final = pd.merge(df_geog_final, df_incautaciones, on='ID_PROCEDIMIENTO', how='left')
            df_detenidos_final = pd.merge(df_geog_final, df_detenidos, on='ID_PROCEDIMIENTO', how='left')
            df_otros_delitos_final = pd.merge(df_geog_final, df_otros_delitos, on='ID_PROCEDIMIENTO', how='right')
            df_trata_final = pd.merge(df_geog_final, df_trata, on='ID_PROCEDIMIENTO', how='right') if len(df_trata) > 0 else pd.DataFrame()

        if self.CONTADOR['ORDEN_SERVICIOS_NUEVOS'] > 0:
            df_afectados_final = pd.merge(df_geog_final, df_afectados, on='ID_PROCEDIMIENTO', how='left',suffixes=('', '_afectados'))
            df_controlados_final = pd.merge(df_geog_final, df_controlados, on='ID_PROCEDIMIENTO', how='left',suffixes=('', '_controlados'))
            df_codigos_final = pd.merge(df_geog_final, df_codigos, on='ID_PROCEDIMIENTO', how='left',suffixes=('', '_codigos'))


        
        
        dataframes = {
                'GEOG. PROCEDIMIENTO': df_geog_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_geog_final' in locals() else pd.DataFrame(),
                'VEHI. Y PERSO. CONTROLADAS': df_controlados_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_controlados_final' in locals() else pd.DataFrame(),
                'PERSONAL Y ELEMENTOS AFECTADOS': df_afectados_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_afectados_final' in locals() else pd.DataFrame(),
                'INCAUTACIONES': df_incautados_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_incautados_final' in locals() else pd.DataFrame(),
                'DETENIDOS Y APREHENDIDOS': df_detenidos_final.reset_index(drop=True).fillna("").replace("", "-")if 'df_detenidos_final' in locals() else pd.DataFrame(),
                'OTROS DELITOS': df_otros_delitos_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_otros_delitos_final' in locals() else pd.DataFrame(),
                'TRATA O TRAFIC PERSONAS': df_trata_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_trata_final' in locals() else pd.DataFrame(),
                'CODIGO OPERATIVO': df_codigos_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_codigos_final' in locals() else pd.DataFrame(),
            }

        return dataframes
    
    def copiar_formato_template(self, dataframes):
        # 1. Cargar template
        wb_template = load_workbook(PATH_TEMPLATE)
        
        # 2. Definir estilos base
        header_style = {
            'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'font': Font(bold=True, color='FFFFFF'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        
        data_style = {
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            ),
            'alignment': Alignment(horizontal='center', vertical='center')
        }

        # 3. Procesar cada hoja
        for sheet_name, df in dataframes.items():
            if sheet_name in wb_template.sheetnames:
                sheet = wb_template[sheet_name]
                template_sheet = wb_template[sheet_name]
                
                # Copiar formatos de columnas
                for column in range(1, template_sheet.max_column + 1):
                    letter = get_column_letter(column)
                    sheet.column_dimensions[letter].width = template_sheet.column_dimensions[letter].width
                
                # Escribir datos con formato
                for i, row in df.iterrows():
                    for j, value in enumerate(row, 1):
                        cell = sheet.cell(row=i+4, column=j+1, value=value)
                        
                        # Aplicar estilo según posición
                        if i == 0:
                            for k, v in header_style.items():
                                setattr(cell, k, v)
                        else:
                            for k, v in data_style.items():
                                setattr(cell, k, v)
                                
                # Preservar fórmulas existentes         
                for row in template_sheet.rows:
                    for cell in row:
                        if cell.value and str(cell.value).startswith('='):
                            sheet[cell.coordinate].value = cell.value

        # 4. Guardar resultado
        PATH_FILE_OUTPUT = f'informes/Informe_{self.fecha_inicial}_AL_{self.fecha_final}.xlsx'
        wb_template.save(PATH_FILE_OUTPUT)
        
        # Abrir el archivo Excel usando el programa predeterminado del sistema
        os.startfile(os.path.abspath(PATH_FILE_OUTPUT))
    
    def actualizar_base_datos(self):
        """Actualiza la base de datos con los nuevos registros"""
        try:
            fecha_actual = datetime.now().strftime('%Y%m%d')

            try:
                wb_base = load_workbook(PATH_BASE)
            except:
                wb_base = Workbook()
                wb_base.remove(wb_base.active)  # Eliminar hoja por defecto
                
                # Crear hojas necesarias
                wb_base.create_sheet("PROCEDIMIENTOS")
                wb_base.create_sheet("DETENIDOS_APREHENDIDOS")
                wb_base.create_sheet("OTROS DELITOS") 
                wb_base.create_sheet("INCAUTACIONES")
                wb_base.create_sheet("TRATA")
                wb_base.save("BASE_DATOS.xlsx")

            # Actualizar cada hoja con los nuevos datos
            hojas = {
                "PROCEDIMIENTOS": self.df_procedimientos,
                "DETENIDOS_APREHENDIDOS": self.df_detenidos,
                "OTROS DELITOS": self.df_otros_delitos,
                "INCAUTACIONES": self.df_incautaciones,
                "TRATA": self.df_trata if hasattr(self, 'df_trata') else None
            }

            for nombre_hoja, df in hojas.items():
                if df is not None and not df.empty:
                    # Convertir DataFrame a filas
                    filas = [df.columns.tolist()] + df.values.tolist()
                    
                    # Limpiar hoja actual
                    hoja = wb_base[nombre_hoja]
                    hoja.delete_rows(1, hoja.max_row)
                    
                    # Escribir nuevos datos
                    for fila in filas:
                        hoja.append(fila)

                    # Ajustar ancho de columnas
                    for columna in hoja.columns:
                        max_length = 0
                        for celda in columna:
                            try:
                                if len(str(celda.value)) > max_length:
                                    max_length = len(str(celda.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        hoja.column_dimensions[columna[0].column_letter].width = adjusted_width

            # Guardar cambios
            wb_base.save("BASE_DATOS.xlsx")
            
            # Hacer backup
            wb_base.save(f"BASE_DATOS_{fecha_actual}.xlsx")

            self.progress.emit(98, "Base de datos actualizada correctamente")
            return True

        except Exception as e:
            raise Exception(f"Error actualizando base de datos: {str(e)}")

    def ordenar_columnas(self, dataframes):
        """Orders columns for each sheet according to template"""
        
        COLUMN_ORDERS = {
            "GEOG. PROCEDIMIENTO": [
                'FUERZA_INTERVINIENTE', 'ID_OPERATIVO', 'ID_PROCEDIMIENTO', 'UNIDAD_INTERVINIENTE',
                'DESCRIPCIÓN', 'TIPO_INTERVENCION', 'PROVINCIA', 'DEPARTAMENTO O PARTIDO',
                'LOCALIDAD', 'DIRECCION', 'ZONA_SEGURIDAD_FRONTERAS', 'PASO_FRONTERIZO',
                'LATITUD', 'LONGITUD', 'FECHA', 'HORA', 'OTRAS AGENCIAS INTERVINIENTES',
                'Observaciones - Detalles'
            ],
            "VEHI. Y PERSO. CONTROLADAS": [
                'FUERZA_INTERVINIENTE', 'ID_OPERATIVO', 'ID_PROCEDIMIENTO', 'UNIDAD_INTERVINIENTE',
                'DESCRIPCIÓN', 'TIPO_INTERVENCION', 'VEHICULOS_CONTROLADOS',
                'PERSONAS_CONTROLADAS', 'CANT_AVERIGUACIONES_SECUESTRO',
                'CANT_SOLICITUDES_ANTECEDENTES'
            ],
            "PERSONAL Y ELEMENTOS AFECTADOS": [
                'FUERZA_INTERVINIENTE', 'ID_OPERATIVO', 'ID_PROCEDIMIENTO', 'UNIDAD_INTERVINIENTE',
                'DESCRIPCIÓN', 'TIPO_INTERVENCION', 'CANT_EFECTIVOS', 'CANT_AUTOS_CAMIONETAS',
                'CANT_SCANNERS', 'CANT_EMBARCACIONES', 'CANT_MOTOS', 'CANT_CABALLOS',
                'CANT_CANES', 'CANT_MORPHRAPID', 'CANT_LPR'
            ],
            "INCAUTACIONES": [
                'FUERZA_INTERVINIENTE', 'ID_OPERATIVO', 'ID_PROCEDIMIENTO', 'UNIDAD_INTERVINIENTE',
                'DESCRIPCIÓN', 'TIPO_INTERVENCION', 'TIPO_INCAUTACION', 'TIPO', 'SUBTIPO',
                'CANTIDAD', 'MEDIDAS', 'AFORO', 'OBSERVACIONES'
            ],
            "DETENIDOS Y APREHENDIDOS": [
                'FUERZA_INTERVINIENTE', 'ID_OPERATIVO', 'ID_PROCEDIMIENTO', 'UNIDAD_INTERVINIENTE',
                'DESCRIPCIÓN', 'TIPO_INTERVENCION', 'EDAD', 'SEXO', 'GENERO', 'NACIONALIDAD',
                'SITUACION_PROCESAL', 'DELITO_IMPUTADO', 'JUZGADO_INTERVINIENTE',
                'CARATULA_CAUSA', 'NUM_CAUSA'
            ],
            "OTROS DELITOS": [
                'FUERZA_INTERVINIENTE', 'ID_OPERATIVO', 'ID_PROCEDIMIENTO', 'UNIDAD_INTERVINIENTE',
                'DESCRIPCIÓN', 'TIPO_INTERVENCION', 'TIPO_OTRO_DELITO', 'GENERO_VICTIMA',
                'EDAD_VICTIMA', 'NACIONALIDAD', 'OBSERVACIONES'
            ],
            "TRATA O TRAFIC PERSONAS": [
                'FUERZA_INTERVINIENTE', 'ID_OPERATIVO', 'ID_PROCEDIMIENTO', 'UNIDAD_INTERVINIENTE',
                'DESCRIPCIÓN', 'TIPO_INTERVENCION', 'TIPO_DELITO', 'SEXO_VICTIMA', 
                'GENERO_VICTIMA', 'EDAD_VICTIMA', 'NACIONALIDAD', 'JUZGADO_INTERVINIENTE',
                'OBSERVACIONES'
            ],
            "CODIGO OPERATIVO": [
                'FUERZA_INTERVINIENTE', 'ID_OPERATIVO', 'ID_PROCEDIMIENTO', 'UNIDAD_INTERVINIENTE',
                'DESCRIPCIÓN', 'TIPO_INTERVENCION', 'CODIGO_OPERATIVO'
            ]
        }
        
        ordered_dfs = {}
    
        try:
            for sheet_name, df in dataframes.items():
                if df is None or df.empty:
                    ordered_dfs[sheet_name] = df
                    continue
                    
                if sheet_name in COLUMN_ORDERS:
                    # Check if all required columns exist
                    missing_cols = set(COLUMN_ORDERS[sheet_name]) - set(df.columns)
                    if missing_cols:
                        print(f"Warning: Missing columns in {sheet_name}: {missing_cols}")
                        # Add missing columns with empty values
                        for col in missing_cols:
                            df[col] = "-"
                    
                    # Reorder only existing columns
                    cols_to_use = [col for col in COLUMN_ORDERS[sheet_name] if col in df.columns]
                    ordered_dfs[sheet_name] = df[cols_to_use]
                else:
                    ordered_dfs[sheet_name] = df
                    
        except Exception as e:
            print(f"Error ordering columns: {str(e)}")
            return dataframes
            
        return ordered_dfs