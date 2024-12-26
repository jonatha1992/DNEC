#!/usr/bin/env python
# coding: utf-8

# # LIBRERIAS 
# 
# * Tener en cuenta bajadas de Procedimientos, Personas Armas, Divisas , vehiculos (secuestrados, ministerio), Narcotrafico general,Objetos y pegar en la carpeta bajadas de fecha del primero del 1 del mes en informar hasta fecha del informe 

# ### IMPORTACIÓN DE LIBREARIAS
# 

# In[1]:


import pandas as pd
from  Parametros import * 
from  Funciones import * 
from openpyxl import load_workbook , Workbook
import os
from datetime import datetime
from openpyxl.styles import Alignment

# Configuración para evitar el FutureWarning
pd.set_option('future.no_silent_downcasting', True)


# ### RUTAS

# In[2]:


PATH_BASE = 'db/base_informada.xlsx'
TEMPLATE_PATH = 'models/Planilla_modelo.xlsx'
PATH_PROCEDIMIENTOS = obtener_ruta_bajada('procedimiento')
PATH_ARMAS = obtener_ruta_bajada('arma')
PATH_DIVISAS = obtener_ruta_bajada('divisa')
PATH_NARCOTRAFICO =  obtener_ruta_bajada('narcotrafico')
PATH_OBJETOS =  obtener_ruta_bajada('objetos')
PATH_PERSONAS =  obtener_ruta_bajada('persona')
PATH_VEHICULOS =  obtener_ruta_bajada('vehiculo')
PATH_OPERACIONES =  obtener_ruta_bajada('OPER')
PATH_TRATA =  obtener_ruta_bajada('trata')


# ### PARAMETROS
# 

# In[3]:


FECHA_MINIMA = '11-12-2024'
FECHA_MAYOR = '17-12-2024'

VERIFICAR_CON_BASE = False
ACTUALIZAR_BASE = False

CONTADOR={
    "OPERATIVOS_BASE":0,
    "PROCEDIMIENTOS_BASE":0,
    "FECHA_MENOR_BASE":datetime,
    "FECHA_MAYOR_BASE":datetime,
    'PROCEDIMIENTOS_NUEVOS': 0,
    'ORDEN_SERVICIOS_NUEVOS': 0,
    'BAJADA_ORDEN_SERVICIOS': 0,
    'BAJADA_PROCEDIMIENTOS': 0,
    'BAJADA_ARMAS': 0,
    'BAJADA_DIVISAS': 0,
    'BAJADA_NARCOTRAFICO': 0,
    'BAJADA_OBJETOS': 0,
    'BAJADA_PERSONAS': 0,
    'BAJADA_VEHICULOS': 0,
    'BAJADA_TRATA': 0,
    'GEOG_FINAL': 0,
    'VICTIMAS_FINAL': 0,
    'DETENIDOS_FINAL': 0,
    'ARMAS_FINAL': 0,
    'DIVISAS_FINAL': 0,
    'NARCOTRAFICO_FINAL': 0,
    'OBJETOS_FINAL': 0,
    'VEHICULOS_FINAL': 0,
    'TRATA_FINAL': 0,
}


# ### ACCEDER BASE DE DATOS

# In[4]:


try:
    # Intentar leer el archivo
    book_base_geo = pd.read_excel(PATH_BASE, sheet_name="GEOG. PROCEDIMIENTO")
    
    print("\nValores únicos por columna:")
    print(book_base_geo.nunique())  
    
    CONTADOR['OPERATIVOS_BASE'] = book_base_geo['ID_OPERATIVO'].nunique()
    CONTADOR['PROCEDIMIENTOS_BASE'] = book_base_geo['ID_PROCEDIMIENTO'].nunique()
    CONTADOR['FECHA_MENOR_BASE'] = book_base_geo['FECHA'].min()
    CONTADOR['FECHA_MAYOR_BASE'] = book_base_geo['FECHA'].max()
    

    print(f"MENOR FECHA DE BASE DATOS: {CONTADOR.get('FECHA_MENOR_BASE')}")
    print(f"MAYOR FECHA DE BASE DATOS: {CONTADOR.get('FECHA_MAYOR_BASE')}")
    
except FileNotFoundError:
    # Manejar el error de archivo no encontrado
    print("El archivo 'base_informada.xlsx' no fue encontrado en la ruta 'db/'. Verifica su existencia.")
    book_base_geo = pd.DataFrame()
    


# # SIGIPOL
# 

# ### PROCEDIMIENTOS GENERALES
# 
# 
# 

# In[5]:


excel_bajada_procedimientos_generales = filtrar_procedimientos_generales(PATH_PROCEDIMIENTOS)
CONTADOR['BAJADA_PROCEDIMIENTOS'] = len(excel_bajada_procedimientos_generales)
df_procedimientos = pd.DataFrame()
df_procedimientos['ID_OPERATIVO'] = excel_bajada_procedimientos_generales.apply(procesar_causa_judicial, axis=1)
df_procedimientos['FUERZA_INTERVINIENTE'] = "PSA"
df_procedimientos['ID_PROCEDIMIENTO'] = excel_bajada_procedimientos_generales.apply(generar_uid_sigpol, axis=1)
df_procedimientos['CAUSAJUDICIALNUMERO'] = excel_bajada_procedimientos_generales['CAUSAJUDICIALNUMERO'].copy()
df_procedimientos['UNIDAD_INTERVINIENTE'] = excel_bajada_procedimientos_generales['UOSP'].fillna(excel_bajada_procedimientos_generales['URSA'])
df_procedimientos['DESCRIPCIÓN'] = excel_bajada_procedimientos_generales.apply(procesar_descripcion, axis=1)
df_procedimientos['TIPO_INTERVENCION'] = excel_bajada_procedimientos_generales.apply(procesar_tipo_procedimiento, axis=1)
df_procedimientos['PROVINCIA'] = excel_bajada_procedimientos_generales.apply(procesar_provincia, axis=1)
df_procedimientos['DEPARTAMENTO O PARTIDO'] = excel_bajada_procedimientos_generales.apply(procesar_municipio, axis=1)
df_procedimientos['LOCALIDAD'] = "-"
df_procedimientos['DIRECCION'] = excel_bajada_procedimientos_generales.apply(procesar_direccion, axis=1)
df_procedimientos[['LATITUD', 'LONGITUD']] = excel_bajada_procedimientos_generales.apply(procesar_geog, axis=1, result_type='expand')
df_procedimientos['FECHA'] = pd.to_datetime(excel_bajada_procedimientos_generales['DENUNCIAFECHA'], errors='coerce').dt.date
df_procedimientos['HORA'] = pd.to_datetime(excel_bajada_procedimientos_generales['DENUNCIAFECHA'], errors='coerce').dt.strftime('%H:%M')
df_procedimientos['ZONA_SEGURIDAD_FRONTERAS'] = "-"
df_procedimientos['PASO_FRONTERIZO'] = "-"
df_procedimientos['OTRAS AGENCIAS INTERVINIENTES'] = "-"
df_procedimientos['Observaciones - Detalles'] = "-"


df_procedimientos['LATITUD'] = df_procedimientos['LATITUD'].astype(str).str.replace(',', '.')
df_procedimientos['LONGITUD'] = df_procedimientos['LONGITUD'].astype(str).str.replace(',', '.')

df_procedimientos_completado = df_procedimientos[['FUERZA_INTERVINIENTE', 'ID_OPERATIVO', 'ID_PROCEDIMIENTO',
                                     'UNIDAD_INTERVINIENTE', 'DESCRIPCIÓN', 'TIPO_INTERVENCION',
                                     'PROVINCIA', 'DEPARTAMENTO O PARTIDO', 'LOCALIDAD', 'DIRECCION',
                                     'ZONA_SEGURIDAD_FRONTERAS', 'PASO_FRONTERIZO', 'LATITUD', 'LONGITUD',
                                     'FECHA', 'HORA', 'OTRAS AGENCIAS INTERVINIENTES', 'Observaciones - Detalles']]

print(df_procedimientos_completado.nunique())


# ### FILTRO CON LA BASE DE DATOS DE LO QUE SE FUE INFORMADO
# 
# 
# 

# In[6]:


if  book_base_geo.empty :
    print("No hay casos en la base de datos ")
    df_procedimientos_final = df_procedimientos_completado
elif  not book_base_geo.empty and not VERIFICAR_CON_BASE :    
    df_procedimientos_final = df_procedimientos_completado
else:
    df_procedimientos_final = df_procedimientos_completado[~df_procedimientos_completado['ID_PROCEDIMIENTO'].isin(set(book_base_geo['ID_PROCEDIMIENTO']))]
    
CONTADOR['PROCEDIMIENTOS_NUEVOS'] = len(df_procedimientos_final['ID_PROCEDIMIENTO'])
print(df_procedimientos_final.nunique())




# ### HOJA PERSONAS  

# In[7]:


# Paso 1: Cargar datos iniciales

if CONTADOR['PROCEDIMIENTOS_NUEVOS'] != 0:
    excel_bajada_personas = pd.read_excel(PATH_PERSONAS)

    if not excel_bajada_personas.empty:
        CONTADOR['BAJADA_PERSONAS'] = len(excel_bajada_personas)

        # Paso 2: Generar UID
        excel_bajada_personas['UOSP'] = excel_bajada_personas['UOSP'].fillna(excel_bajada_personas['URSA'])
        excel_bajada_personas['TIPO_CAUSA_INTERNA'] = excel_bajada_personas.apply( procesar_tipo_causa_interna, axis=1)
        excel_bajada_personas['ID_PROCEDIMIENTO'] = excel_bajada_personas.apply(generar_uid_sigpol, axis=1)

        # Paso 3: Coincidencia de UID entre excel_bajada_personas y df_geog_final
        uid_coincidentes = excel_bajada_personas[excel_bajada_personas['ID_PROCEDIMIENTO'].isin(df_procedimientos_final['ID_PROCEDIMIENTO'])]

        # Crear DataFrame de detenidos y aprendidos
        df_detenidos_aprendidos = pd.DataFrame()
        df_detenidos_aprendidos['ID_PROCEDIMIENTO'] = uid_coincidentes['ID_PROCEDIMIENTO']
        df_detenidos_aprendidos['EDAD'] = uid_coincidentes.apply(procesar_edad, axis=1)
        df_detenidos_aprendidos['SEXO'] = uid_coincidentes.apply(procesar_sexo, axis=1)
        df_detenidos_aprendidos['GENERO'] = uid_coincidentes.apply(procesar_genero, axis=1)
        df_detenidos_aprendidos['NACIONALIDAD'] = uid_coincidentes.apply(procesar_nacionalidad, axis=1)
        df_detenidos_aprendidos['SITUACION_PROCESAL'] = uid_coincidentes.apply(procesar_situacion_judicial, axis=1)
        df_detenidos_aprendidos['DELITO_IMPUTADO'] = uid_coincidentes.apply(procesar_tipo_delito, axis=1)
        df_detenidos_aprendidos['JUZGADO_INTERVINIENTE'] = uid_coincidentes.apply(procesar_juzgado, axis=1)
        df_detenidos_aprendidos['CARATULA_CAUSA'] = uid_coincidentes.apply(procesar_caratula, axis=1)
        df_detenidos_aprendidos['NUM_CAUSA'] = uid_coincidentes['ID_PROCEDIMIENTO']


        df_detenidos_aprendidos = df_detenidos_aprendidos[df_detenidos_aprendidos['SITUACION_PROCESAL'] != 'NO INFORMAR']
        df_detenidos_aprendidos = df_detenidos_aprendidos[df_detenidos_aprendidos['ID_PROCEDIMIENTO'].isin(df_procedimientos_final['ID_PROCEDIMIENTO'])]
        print(df_detenidos_aprendidos.nunique())


# ### HOJA OTROS DELITOS

# In[8]:


if CONTADOR['PROCEDIMIENTOS_NUEVOS'] != 0:
    
    if not excel_bajada_personas.empty:
        
        df_otros_delitos = df_detenidos_aprendidos[df_detenidos_aprendidos['SITUACION_PROCESAL'] == 'VICTIMA']
        # Filtro los que son victima de detenidos y aprendidos
        df_detenidos_aprendidos_completado = df_detenidos_aprendidos[df_detenidos_aprendidos['SITUACION_PROCESAL'] != 'VICTIMA']

        # Cambiar el nombre de algunas columnas en el DataFrame de víctimas
        df_otros_delitos_completado = df_otros_delitos.rename(columns={
            'DELITO_IMPUTADO': 'TIPO_OTRO_DELITO',
            'GENERO': 'GENERO_VICTIMA',
            'EDAD': 'EDAD_VICTIMA',
        }) # type: ignore
        df_otros_delitos_completado["OBSERVACIONES"] ="-"
        
        CONTADOR['DETENIDOS_FINAL'] = len(df_detenidos_aprendidos_completado['ID_PROCEDIMIENTO'])
        CONTADOR['VICTIMAS_FINAL'] = len(df_otros_delitos_completado['ID_PROCEDIMIENTO'])
        print(df_otros_delitos_completado.nunique())


# ### INCAUTACIONES ARMA

# In[9]:


if CONTADOR['PROCEDIMIENTOS_NUEVOS'] != 0:

    # Paso 1: Cargar datos iniciales
    excel_bajada_arma = pd.read_excel(PATH_ARMAS)
    
    if not excel_bajada_arma.empty:
        CONTADOR['BAJADA_ARMAS'] = len(excel_bajada_arma)

        excel_bajada_arma = excel_bajada_arma[excel_bajada_arma["TIPO_ESTADO_OBJETO"] == "SECUESTRADO"]
        excel_bajada_arma['UOSP'] = excel_bajada_arma['UOSP'].fillna(excel_bajada_arma['URSA'])
        excel_bajada_arma['TIPO_CAUSA_INTERNA'] = excel_bajada_arma.apply( procesar_tipo_causa_interna, axis=1)
        excel_bajada_arma['ID_PROCEDIMIENTO'] = excel_bajada_arma.apply(generar_uid_sigpol, axis=1)


        df_arma = pd.DataFrame()
        df_arma['ID_PROCEDIMIENTO'] = excel_bajada_arma['ID_PROCEDIMIENTO']
        df_arma['TIPO_INCAUTACION'] = "ARMAS"
        df_arma['TIPO'] = excel_bajada_arma['TIPO_ARMA']
        df_arma['SUBTIPO'] = "-"
        df_arma['CANTIDAD'] = excel_bajada_arma.apply(procesar_cantidad_arma, axis=1)
        df_arma['MEDIDAS'] = "UNIDADES"
        df_arma['AFORO'] = "-"
        df_arma['OBSERVACIONES'] = excel_bajada_arma.apply(procesar_observaciones_arma, axis=1)

        df_arma_completado = df_arma[df_arma['ID_PROCEDIMIENTO'].isin(df_procedimientos_final['ID_PROCEDIMIENTO'])]

        CONTADOR['ARMAS_FINAL'] = len(df_arma['ID_PROCEDIMIENTO'])
        print(df_arma_completado.nunique())


# ### INCAUTACIONES DIVISA

# In[10]:


if CONTADOR['PROCEDIMIENTOS_NUEVOS'] != 0:
    
    excel_bajada_divisa = pd.read_excel(PATH_DIVISAS)
    
    if not excel_bajada_divisa.empty:

        CONTADOR['BAJADA_DIVISAS'] = len(excel_bajada_divisa)
        excel_bajada_procedimientos_generales['PARTE_ANIO'] = excel_bajada_procedimientos_generales['NUMERO_PARTE'].astype(str) + "/" + excel_bajada_procedimientos_generales['ANIO_PARTE'].astype(str)
        excel_bajada_divisa['PARTE_ANIO'] = excel_bajada_divisa['NUMERO_PARTE'].astype(str) + "/" + excel_bajada_divisa['ANIO_PARTE'].astype(str)

        excel_bajada_divisa['UOSP'] = excel_bajada_divisa['UOSP'].fillna(excel_bajada_divisa['URSA'])

        excel_bajada_divisa = pd.merge(excel_bajada_divisa, 
                                    excel_bajada_procedimientos_generales[['PARTE_ANIO', 'TIPO_CAUSA_INTERNA']], 
                                    on='PARTE_ANIO', 
                                    how='left')

        excel_bajada_divisa['ID_PROCEDIMIENTO'] = excel_bajada_divisa.apply(generar_uid_sigpol, axis=1)
        excel_bajada_divisa = excel_bajada_divisa[~pd.isnull(excel_bajada_divisa['TOTAL_DIVISAS_SECUESTRADAS'])]


        cantidad_por_uid = excel_bajada_divisa.groupby(['ID_PROCEDIMIENTO', 'TIPO_DIVISA'], as_index=False)['TOTAL_DIVISAS_SECUESTRADAS'].sum()

        df_divisa = pd.DataFrame()
        df_divisa['ID_PROCEDIMIENTO'] = cantidad_por_uid['ID_PROCEDIMIENTO']
        df_divisa['TIPO_INCAUTACION'] = "DIVISAS"
        df_divisa['TIPO'] = cantidad_por_uid['TIPO_DIVISA']
        df_divisa['SUBTIPO'] = "-"
        df_divisa['CANTIDAD'] = cantidad_por_uid['TOTAL_DIVISAS_SECUESTRADAS']
        df_divisa['MEDIDAS'] = "UNIDADES"
        df_divisa['AFORO'] = "-"
        df_divisa['OBSERVACIONES'] = "-"

        df_divisa_completado = df_divisa[df_divisa['ID_PROCEDIMIENTO'].isin(df_procedimientos_final['ID_PROCEDIMIENTO'])]
        
        CONTADOR['DIVISAS_FINAL'] = len(df_divisa_completado['ID_PROCEDIMIENTO'])
        print(df_divisa_completado.nunique())


# ### INCAUTACIONES OBJETOS

# In[11]:


if CONTADOR['PROCEDIMIENTOS_NUEVOS'] != 0:

    excel_bajada_objetos = pd.read_excel(PATH_OBJETOS)
    CONTADOR ['BAJADA_OBJETOS'] = len(excel_bajada_objetos)
    excel_bajada_objetos = excel_bajada_objetos[excel_bajada_objetos["ESTADO"] == "SECUESTRADO"]

    excel_bajada_objetos['UOSP'] = excel_bajada_objetos['UOSP'].fillna(excel_bajada_objetos['URSA'])
    excel_bajada_objetos['TIPO_CAUSA_INTERNA'] = excel_bajada_objetos.apply( procesar_tipo_causa_interna, axis=1)
    excel_bajada_objetos['ID_PROCEDIMIENTO'] = excel_bajada_objetos.apply(generar_uid_sigpol, axis=1)
    excel_bajada_objetos = excel_bajada_objetos[excel_bajada_objetos['TIPO_OBJETO'].isin(TIPO_OBJETO)]
    excel_bajada_objetos = excel_bajada_objetos[~pd.isnull(excel_bajada_objetos['CANTIDAD'])]

    df_objetos = pd.DataFrame()
    df_objetos['ID_PROCEDIMIENTO'] = excel_bajada_objetos['ID_PROCEDIMIENTO']
    df_objetos['TIPO_INCAUTACION'] = "MERCADERIA"
    df_objetos['TIPO'] = excel_bajada_objetos['TIPO_OBJETO']
    df_objetos['SUBTIPO'] = "-"
    df_objetos['CANTIDAD'] = excel_bajada_objetos['CANTIDAD']
    df_objetos['MEDIDAS'] = "UNIDADES"
    df_objetos['AFORO'] = "-"
    df_objetos['OBSERVACIONES'] = "-"

    df_objetos_completado = df_objetos[df_objetos['ID_PROCEDIMIENTO'].isin(df_procedimientos_final['ID_PROCEDIMIENTO'])]

    CONTADOR['OBJETOS_FINAL'] = len(df_objetos_completado['ID_PROCEDIMIENTO'])
    print(df_objetos_completado.nunique())


# ### INCAUTACIONES VEHICULOS

# In[12]:


if CONTADOR['PROCEDIMIENTOS_NUEVOS'] != 0:
    
    excel_bajada_vehiculos = pd.read_excel(PATH_VEHICULOS)
    
    if not excel_bajada_vehiculos.empty:
        CONTADOR ['BAJADA_VEHICULOS'] = len(excel_bajada_vehiculos)

        excel_bajada_vehiculos = excel_bajada_vehiculos[excel_bajada_vehiculos["VEHICULO_ESTADO"] == "SECUESTRADO"]
        excel_bajada_vehiculos['TIPO_CAUSA_INTERNA'] = excel_bajada_vehiculos.apply( procesar_tipo_causa_interna, axis=1)
        excel_bajada_vehiculos['UOSP'] = excel_bajada_vehiculos['UOSP'].fillna(excel_bajada_vehiculos['URSA'])
        excel_bajada_vehiculos['TIPO_CAUSA_INTERNA'] = excel_bajada_vehiculos.apply( procesar_tipo_causa_interna, axis=1)
        excel_bajada_vehiculos['ID_PROCEDIMIENTO'] = excel_bajada_vehiculos.apply(generar_uid_sigpol, axis=1)


        df_vehiculos = pd.DataFrame()
        df_vehiculos['ID_PROCEDIMIENTO'] = excel_bajada_vehiculos['ID_PROCEDIMIENTO']
        df_vehiculos['TIPO_INCAUTACION'] = "SECUESTRO DE VEHICULOS"
        df_vehiculos['TIPO'] = excel_bajada_vehiculos['VEHICULO_TIPO']
        df_vehiculos['SUBTIPO'] = "-"
        df_vehiculos['CANTIDAD'] = excel_bajada_vehiculos['CANTIDAD']
        df_vehiculos['MEDIDAS'] = "UNIDADES"
        df_vehiculos['AFORO'] = "-"
        df_vehiculos['OBSERVACIONES'] = excel_bajada_vehiculos.apply(observaciones_vehiculo,axis=1)


        df_vehiculos_completado = df_vehiculos[df_vehiculos['ID_PROCEDIMIENTO'].isin(df_procedimientos_final['ID_PROCEDIMIENTO'])]
        
        CONTADOR['VEHICULOS_FINAL'] = len(df_vehiculos_completado['ID_PROCEDIMIENTO'])
        print(df_vehiculos_completado.nunique())


# ### INCAUTACIONES NARCOTRAFICO

# In[13]:


if CONTADOR['PROCEDIMIENTOS_NUEVOS'] != 0:
    
    # Paso 1: Cargar datos iniciales
    excel_bajada_narcotrafico = pd.read_excel(PATH_NARCOTRAFICO)
    if not excel_bajada_narcotrafico.empty:
        
        CONTADOR['BAJADA_NARCOTRAFICO'] = len(excel_bajada_narcotrafico)
        
        excel_bajada_narcotrafico['UOSP'] = excel_bajada_narcotrafico['UOSP'].fillna(excel_bajada_narcotrafico['URSA'])
        excel_bajada_narcotrafico['TIPO_CAUSA_INTERNA'] = excel_bajada_narcotrafico.apply( procesar_tipo_causa_interna, axis=1)
        excel_bajada_narcotrafico['ID_PROCEDIMIENTO'] = excel_bajada_narcotrafico.apply(generar_uid_sigpol, axis=1)

        excel_bajada_narcotrafico = excel_bajada_narcotrafico[~pd.isnull(excel_bajada_narcotrafico['TIPO_ESTUPEFACIENTE'])]


        #Paso 2: Crear df_divisa con las cantidades sumadas
        df_narcotrafico = pd.DataFrame()
        df_narcotrafico['ID_PROCEDIMIENTO'] = excel_bajada_narcotrafico['ID_PROCEDIMIENTO']
        df_narcotrafico['TIPO_INCAUTACION'] = "ESTUPEFACIENTE"
        df_narcotrafico['TIPO'] = excel_bajada_narcotrafico.apply(clasificar_tipo_sustancia, axis=1)
        df_narcotrafico['SUBTIPO'] = "-"
        df_narcotrafico[['CANTIDAD', 'MEDIDAS']] = excel_bajada_narcotrafico.apply(clasificar_medida, axis=1, result_type='expand')
        df_narcotrafico['AFORO'] = "-"
        df_narcotrafico['OBSERVACIONES'] = excel_bajada_narcotrafico.apply(observaciones_sustancia, axis=1)

        df_narcotrafico_completado = df_narcotrafico[df_narcotrafico['ID_PROCEDIMIENTO'].isin(df_procedimientos_final['ID_PROCEDIMIENTO'])]
        
        CONTADOR['NARCOTRAFICO_FINAL'] = len(df_narcotrafico_completado['ID_PROCEDIMIENTO'])
        print(df_narcotrafico_completado.nunique())


# ### UNION DE INCAUTACIONES 
# ARMAS, DIVISAS, OBJETOS , VEHICULOS , NARCOTRAFICO

# In[14]:


if CONTADOR['PROCEDIMIENTOS_NUEVOS'] != 0:
    df_incautados_completado= pd.concat([df_objetos_completado, df_vehiculos_completado, df_arma_completado, df_divisa_completado , df_narcotrafico_completado])
    print(df_incautados_completado.nunique())


# ### TRATA DE PERSONA

# In[15]:


if CONTADOR['PROCEDIMIENTOS_NUEVOS'] != 0:

    excel_bajada_trata = pd.read_excel(PATH_TRATA)
    if not excel_bajada_trata.empty:
        CONTADOR['BAJADA_TRATA'] = len(excel_bajada_trata)

        excel_bajada_trata['UOSP'] = excel_bajada_trata['UOSP'].fillna(excel_bajada_trata['URSA'])
        excel_bajada_trata['TIPO_CAUSA_INTERNA'] = excel_bajada_trata.apply( procesar_tipo_causa_interna, axis=1)
        excel_bajada_trata['ID_PROCEDIMIENTO'] = excel_bajada_trata.apply(generar_uid_sigpol, axis=1)

        # Crear DataFrame de detenidos y aprendidos
        df_trata = pd.DataFrame()
        df_trata['ID_PROCEDIMIENTO'] = excel_bajada_trata['ID_PROCEDIMIENTO']
        df_trata['TIPO_DELITO'] = excel_bajada_trata["TIPO_EXPLOTACION"].fillna("-")
        df_trata['SEXO_VICTIMA'] = excel_bajada_trata.apply(procesar_sexo, axis=1)
        df_trata['GENERO_VICTIMA'] = excel_bajada_trata.apply(procesar_genero, axis=1)
        df_trata['EDAD_VICTIMA'] = excel_bajada_trata['EDAD'].fillna("-")
        df_trata['NACIONALIDAD'] = excel_bajada_trata['NACIONALIDAD'].fillna("-")
        df_trata['NACIONALIDAD'] = excel_bajada_trata['NACIONALIDAD'].str.upper()
        df_trata['JUZGADO_INTERVINIENTE'] = excel_bajada_trata.apply(procesar_juzgado, axis=1)
        df_trata['OBSERVACIONES'] = "-"

        df_trata_completado = df_trata[df_trata['ID_PROCEDIMIENTO'].isin(df_procedimientos_final['ID_PROCEDIMIENTO'])]
        df_trata_completado = df_trata[df_trata['TIPO_DELITO'] != "-"]
        
        CONTADOR['TRATA_FINAL'] = len(df_trata_completado['ID_PROCEDIMIENTO'])
        print(df_trata_completado.nunique())



# # OPERACIONES

# ### ORDEN DE SERVICIOS

# In[16]:


excel_bajada_operaciones = pd.read_excel(PATH_OPERACIONES, 
                                         sheet_name="ORDEN_SERVICIOS", 
                                         skiprows=1)

CONTADOR['BAJADA_ORDEN_SERVICIOS'] = len(excel_bajada_operaciones)
# Usar .strftime('%H:%M') en cada valor para obtener solo la hora y el minuto
excel_bajada_operaciones['HORA'] = excel_bajada_operaciones['HORA'].apply(lambda x: x.strftime('%H:%M') if pd.notnull(x) else None)
excel_bajada_operaciones["PROVINCIA"] = excel_bajada_operaciones["PROVINCIA"].astype(str).str.strip()
# Ahora copiamos esta columna al nuevo DataFrame

print(excel_bajada_operaciones["PROVINCIA"].str.contains("_").sum())

df_operaciones= pd.DataFrame()
df_operaciones["ID_PROCEDIMIENTO"] = excel_bajada_operaciones["ID_PROCEDIMIENTO"]
df_operaciones["FUERZA_INTERVINIENTE"] = "PSA"
df_operaciones["ID_OPERATIVO"] = excel_bajada_operaciones["ID_OPERATIVO"]
df_operaciones["UNIDAD_INTERVINIENTE"] = excel_bajada_operaciones["UNIDAD_INTERVINIENTE"]
df_operaciones["DESCRIPCIÓN"] = excel_bajada_operaciones["DESCRIPCIÓN"]
df_operaciones["TIPO_INTERVENCION"] = excel_bajada_operaciones["TIPO_INTERVENCION"]
df_operaciones["PROVINCIA"] = excel_bajada_operaciones["PROVINCIA"].str.replace("_", " ", regex=False)
df_operaciones["DEPARTAMENTO O PARTIDO"] = excel_bajada_operaciones["DEPARTAMENTO O PARTIDO"].str.upper()
df_operaciones["LOCALIDAD"] = excel_bajada_operaciones["LOCALIDAD"]
df_operaciones["DIRECCION"] = excel_bajada_operaciones["DIRECCION"]
df_operaciones['FECHA'] = pd.to_datetime(excel_bajada_operaciones["FECHA"]).dt.date
df_operaciones['HORA'] = excel_bajada_operaciones['HORA']
df_operaciones["ZONA_SEGURIDAD_FRONTERAS"] = "-"
df_operaciones["PASO_FRONTERIZO"] = "-"
df_operaciones['OTRAS AGENCIAS INTERVINIENTES'] =  excel_bajada_operaciones["OTRAS AGENCIAS INTERVINIENTES"]
df_operaciones['Observaciones - Detalles'] = "PATRULLAJE DINAMICO"
df_operaciones[['LATITUD', 'LONGITUD']] = excel_bajada_operaciones.apply(procesar_geog_oper, axis=1, result_type='expand')


df_operaciones_completado = df_operaciones[['FUERZA_INTERVINIENTE', 'ID_OPERATIVO', 'ID_PROCEDIMIENTO',
                                     'UNIDAD_INTERVINIENTE', 'DESCRIPCIÓN', 'TIPO_INTERVENCION',
                                     'PROVINCIA', 'DEPARTAMENTO O PARTIDO', 'LOCALIDAD', 'DIRECCION',
                                     'ZONA_SEGURIDAD_FRONTERAS', 'PASO_FRONTERIZO', 'LATITUD', 'LONGITUD',
                                     'FECHA', 'HORA', 'OTRAS AGENCIAS INTERVINIENTES', 'Observaciones - Detalles']]



# Reemplazar "S/D" y "N/C" por "-"
df_operaciones_completado.replace(["S/D", "N/C"], "-", inplace=True)

# Reemplazar los valores vacíos (NaN) por "-"
df_operaciones_completado.fillna("-", inplace=True)


print(df_operaciones_completado.nunique())


# ### FILTRO CON LA BASE DE DATOS DE LO QUE SE FUE INFORMADO
# 
# 
# 

# In[17]:


if  book_base_geo.empty:
    print("No hay casos en la base de datos ")
    df_operaciones_final = df_operaciones_completado
elif  not book_base_geo.empty and not VERIFICAR_CON_BASE :    
    df_operaciones_final = df_operaciones_completado
else:
    df_operaciones_final = df_operaciones_completado[~df_operaciones_completado['ID_PROCEDIMIENTO'].isin(set(book_base_geo['ID_PROCEDIMIENTO']))]
    
CONTADOR['ORDEN_SERVICIOS_NUEVOS'] = len(df_operaciones_final['ID_PROCEDIMIENTO'])
print(df_operaciones_final.nunique())




# ### HOJA DE VEHI. Y PERSO. CONTROLADAS

# In[18]:


if CONTADOR['ORDEN_SERVICIOS_NUEVOS'] != 0:

    df_controlados_controlado = pd.DataFrame()
    df_controlados_controlado["UID"] = excel_bajada_operaciones["ID_PROCEDIMIENTO"]
    df_controlados_controlado["FUERZA_INTERVINIENTE"] = "PSA"
    df_controlados_controlado["ID_PROCEDIMIENTO"] = excel_bajada_operaciones["ID_PROCEDIMIENTO"]
    df_controlados_controlado["ID_OPERATIVO"] = excel_bajada_operaciones["ID_OPERATIVO"]
    df_controlados_controlado["UNIDAD_INTERVINIENTE"] = excel_bajada_operaciones["UNIDAD_INTERVINIENTE"]
    df_controlados_controlado["DESCRIPCIÓN"] = excel_bajada_operaciones["DESCRIPCIÓN"]
    df_controlados_controlado["TIPO_INTERVENCION"] = excel_bajada_operaciones["TIPO_INTERVENCION"]
    df_controlados_controlado["VEHICULOS_CONTROLADOS"] = excel_bajada_operaciones["VEHICULOS_CONTROLADOS"]
    df_controlados_controlado["PERSONAS_CONTROLADAS"] = excel_bajada_operaciones["PERSONAS_CONTROLADAS"]
    df_controlados_controlado["CANT_AVERIGUACIONES_SECUESTRO"] = excel_bajada_operaciones["CANT_AVERIGUACIONES_SECUESTRO"]
    df_controlados_controlado["CANT_SOLICITUDES_ANTECEDENTES"] = excel_bajada_operaciones["CANT_SOLICITUDES_ANTECEDENTES"]


    # Reemplazar "S/D" y "N/C" por "-"
    df_controlados_controlado.replace(["S/D", "N/C"], "-", inplace=True)

    # Reemplazar los valores vacíos (NaN) por "-"
    df_controlados_controlado.fillna("-", inplace=True)

    print(df_controlados_controlado.nunique())


# ### PERSONAL Y ELEMENTOS AFECTADOS

# In[19]:


if CONTADOR['ORDEN_SERVICIOS_NUEVOS'] != 0:

    df_afectados_completado = pd.DataFrame()


    df_afectados_completado["UID"] = excel_bajada_operaciones["ID_PROCEDIMIENTO"]
    df_afectados_completado["FUERZA_INTERVINIENTE"] = "PSA"
    df_afectados_completado["ID_PROCEDIMIENTO"] = excel_bajada_operaciones["ID_PROCEDIMIENTO"]
    df_afectados_completado["ID_OPERATIVO"] = excel_bajada_operaciones["ID_OPERATIVO"]
    df_afectados_completado["UNIDAD_INTERVINIENTE"] = excel_bajada_operaciones["UNIDAD_INTERVINIENTE"]
    df_afectados_completado["DESCRIPCIÓN"] = excel_bajada_operaciones["DESCRIPCIÓN"]
    df_afectados_completado["TIPO_INTERVENCION"] = excel_bajada_operaciones["TIPO_INTERVENCION"]
    df_afectados_completado["CANT_EFECTIVOS"] = excel_bajada_operaciones["CANT_EFECTIVOS"]
    df_afectados_completado["CANT_AUTOS_CAMIONETAS"] = excel_bajada_operaciones["CANT_AUTOS_CAMIONETAS"]
    df_afectados_completado["CANT_SCANNERS"] = excel_bajada_operaciones["CANT_SCANNERS"]
    df_afectados_completado["CANT_EMBARCACIONES"] = excel_bajada_operaciones["CANT_EMBARCACIONES"]
    df_afectados_completado["CANT_MOTOS"] = excel_bajada_operaciones["CANT_MOTOS"]
    df_afectados_completado["CANT_CABALLOS"] = excel_bajada_operaciones["CANT_CABALLOS"]
    df_afectados_completado["CANT_CANES"] = excel_bajada_operaciones["CANT_CANES"]
    df_afectados_completado["CANT_MORPHRAPID"] = excel_bajada_operaciones["CANT_MORPHRAPID"]
    df_afectados_completado["CANT_LPR"] = excel_bajada_operaciones["CANT_LPR"]


    # Reemplazar "S/D" y "N/C" por "-"
    df_afectados_completado.replace(["S/D", "N/C"], "-", inplace=True)

    # Reemplazar los valores vacíos (NaN) por "-"
    df_afectados_completado.fillna("-", inplace=True)

    print(df_afectados_completado.nunique())


# ### CODIGO OPERATIVO

# In[20]:


if CONTADOR['ORDEN_SERVICIOS_NUEVOS'] != 0:
    df_codigo_operativo_final = pd.DataFrame()
    df_codigo_operativo_final["UID"] = excel_bajada_operaciones["ID_PROCEDIMIENTO"]
    df_codigo_operativo_final["FUERZA_INTERVINIENTE"] = "PSA"
    df_codigo_operativo_final["ID_PROCEDIMIENTO"] = excel_bajada_operaciones["ID_PROCEDIMIENTO"]
    df_codigo_operativo_final["ID_OPERATIVO"] = excel_bajada_operaciones["ID_OPERATIVO"]
    df_codigo_operativo_final["UNIDAD_INTERVINIENTE"] = excel_bajada_operaciones["UNIDAD_INTERVINIENTE"]
    df_codigo_operativo_final["DESCRIPCIÓN"] = excel_bajada_operaciones["DESCRIPCIÓN"]
    df_codigo_operativo_final["TIPO_INTERVENCION"] = excel_bajada_operaciones["TIPO_INTERVENCION"]
    df_codigo_operativo_final["CODIGO_OPERATIVO"] = excel_bajada_operaciones["CODIGO_OPERATIVO"]


    # Reemplazar "S/D" y "N/C" por "-"
    df_codigo_operativo_final.replace(["S/D", "N/C"], "-", inplace=True)

    # Reemplazar los valores vacíos (NaN) por "-"
    df_codigo_operativo_final.fillna("-", inplace=True)


    print(df_codigo_operativo_final.nunique())


# # UNION DE DE DATAFRAMES 

# ### CONCATENACION DE PROCEDIMIENTOS Y ORDEN DE SERVICIOS , TAMBIEN SE UNEN LAS 

# In[21]:


df_geog_final = pd.concat ([df_procedimientos_final, df_operaciones_final]) 


CONTADOR["GEOG_FINAL"] = len(df_geog_final['ID_PROCEDIMIENTO'])
print(df_geog_final.nunique())


# ### COMPLETO LA INFORMACION DE TRATA Y OTROS DELITOS DADO QUE "TENGO" QUE PEGAR TODA LA INFORMACION DE LA GEOG FINAL

# In[22]:


# *TIENE QUE TRAER LA INFORMACION DE GEOG FINAL Y COMPLETAR LA INFORMACION 
if CONTADOR['PROCEDIMIENTOS_NUEVOS'] != 0:
    
    df_trata_final = pd.merge(df_geog_final, df_trata_completado, on='ID_PROCEDIMIENTO', how='right')
    print("OTROS TRATA FINAL")

    df_trata_final = df_trata_final[[
        'FUERZA_INTERVINIENTE',
        'ID_OPERATIVO',
        'ID_PROCEDIMIENTO',
        'UNIDAD_INTERVINIENTE',
        'DESCRIPCIÓN',
        'TIPO_INTERVENCION',
        "TIPO_DELITO",
        'SEXO_VICTIMA',
        'GENERO_VICTIMA',
        'EDAD_VICTIMA',
        'NACIONALIDAD',
        'JUZGADO_INTERVINIENTE',
        'OBSERVACIONES'
    ]]
    print(df_trata_final.nunique())

    df_otros_delitos_final = pd.merge(df_geog_final, df_otros_delitos_completado, on='ID_PROCEDIMIENTO', how='right')
    print("\nOTROS DELITOS FINAL")

    df_otros_delitos_final = df_otros_delitos_final[[
        'FUERZA_INTERVINIENTE',
        'ID_OPERATIVO',
        'ID_PROCEDIMIENTO',
        'UNIDAD_INTERVINIENTE',
        'DESCRIPCIÓN',
        'TIPO_INTERVENCION',
        "TIPO_OTRO_DELITO",
        'GENERO_VICTIMA',
        'EDAD_VICTIMA',
        'NACIONALIDAD',
        'OBSERVACIONES'
    ]]

    print(df_otros_delitos_final.nunique())


# ### COMPLETO LA INFORMACION DE TRATA Y OTROS DELITOS DADO QUE NO TENGO QUE PEGAR TODA LA INFORMACION DE LA GEOG FINAL

# In[23]:


# *TIENE QUE PEGAR TODAS LAS FILAS DE  GEOG AUNQUE NO TENGA INFORMACION ES DECIR TIENE QUE SER MAYOR A LA HOJA GEO
if CONTADOR['PROCEDIMIENTOS_NUEVOS'] != 0:

    df_incautados_final= pd.merge(df_geog_final, df_incautados_completado, on='ID_PROCEDIMIENTO', how='left')
    df_incautados_final = df_incautados_final[[
        'FUERZA_INTERVINIENTE',
        'ID_OPERATIVO',
        'ID_PROCEDIMIENTO',
        'UNIDAD_INTERVINIENTE',
        'DESCRIPCIÓN',
        'TIPO_INTERVENCION',
        'TIPO_INCAUTACION',
        'TIPO',
        'SUBTIPO',
        'CANTIDAD',
        'MEDIDAS',
        "AFORO",
        'OBSERVACIONES'
    ]]


    print("\nINCAUTADOS FINAL")
    print(df_incautados_final.nunique())

    df_detenidos_aprendidos_final = pd.merge(df_geog_final, df_detenidos_aprendidos_completado, on='ID_PROCEDIMIENTO', how='left')

    print("\nDETENIDOS FINAL")
    df_detenidos_aprendidos_final = df_detenidos_aprendidos_final[[
        'FUERZA_INTERVINIENTE',
        'ID_OPERATIVO',
        'ID_PROCEDIMIENTO',
        'UNIDAD_INTERVINIENTE',
        'DESCRIPCIÓN',
        'TIPO_INTERVENCION',
        'EDAD',
        'SEXO',
        'GENERO',
        'NACIONALIDAD',
        'SITUACION_PROCESAL',
        "DELITO_IMPUTADO",
        'JUZGADO_INTERVINIENTE',
        'CARATULA_CAUSA',
        "NUM_CAUSA"
    ]]
    print(df_detenidos_aprendidos_final.nunique())


if CONTADOR['ORDEN_SERVICIOS_NUEVOS'] != 0:

    print("\nAFECTADOS FINAL")
    df_afectados_final = pd.merge(df_geog_final, df_afectados_completado, on='ID_PROCEDIMIENTO', how='left',     suffixes=('', '_afectados')  # Evitar sufijos por defecto
    )
    df_afectados_final = df_afectados_final[[
        'FUERZA_INTERVINIENTE',
        'ID_OPERATIVO',
        'ID_PROCEDIMIENTO',
        'UNIDAD_INTERVINIENTE',
        'DESCRIPCIÓN',
        'TIPO_INTERVENCION',
        'CANT_EFECTIVOS',
        'CANT_AUTOS_CAMIONETAS',
        'CANT_SCANNERS',
        'CANT_EMBARCACIONES',
        'CANT_MOTOS',
        "CANT_CABALLOS",
        'CANT_CANES',
        'CANT_MORPHRAPID',
        "CANT_LPR"
    ]]

    print(df_afectados_final.nunique())


    df_controlados_final = pd.merge(df_geog_final, df_controlados_controlado, on='ID_PROCEDIMIENTO', how='left', suffixes=('', '_controlados'))
    print("\nCONTROLADOS FINAL")

    df_controlados_final = df_controlados_final[[
        'FUERZA_INTERVINIENTE',
        'ID_OPERATIVO',
        'ID_PROCEDIMIENTO',
        'UNIDAD_INTERVINIENTE',
        'DESCRIPCIÓN',
        'TIPO_INTERVENCION',
        'VEHICULOS_CONTROLADOS',
        'PERSONAS_CONTROLADAS',
        "CANT_AVERIGUACIONES_SECUESTRO",
        'CANT_SOLICITUDES_ANTECEDENTES',
    ]]

    print(df_controlados_final.nunique())


    df_codigo_operativo_final = pd.merge(df_geog_final, df_codigo_operativo_final, on='ID_PROCEDIMIENTO', how='left',suffixes=('', '_codigos'))
    print("\nCODIGO FINAL")
    df_codigo_operativo_final = df_codigo_operativo_final[[
        'FUERZA_INTERVINIENTE',
        'ID_OPERATIVO',
        'ID_PROCEDIMIENTO',
        'UNIDAD_INTERVINIENTE',
        'DESCRIPCIÓN',
        'TIPO_INTERVENCION',
        'CODIGO_OPERATIVO',
    ]]
    print(df_codigo_operativo_final.nunique())



# # GENERACIÓN DE INFORME

#  ### CREACION DE INFORME SEMANAL CON FECHA HOY

# In[24]:


dataframes = {
    'GEOG. PROCEDIMIENTO': df_geog_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_geog_final' in locals() else pd.DataFrame(),
    'VEHI. Y PERSO. CONTROLADAS': df_controlados_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_controlados_final' in locals() else pd.DataFrame(),
    'PERSONAL Y ELEMENTOS AFECTADOS': df_afectados_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_afectados_final' in locals() else pd.DataFrame(),
    'INCAUTACIONES': df_incautados_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_incautados_final' in locals() else pd.DataFrame(),
    'DETENIDOS Y APREHENDIDOS': df_detenidos_aprendidos_final.reset_index(drop=True).fillna("").replace("", "-")if 'df_detenidos_aprendidos_final' in locals() else pd.DataFrame(),
    'OTROS DELITOS': df_otros_delitos_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_otros_delitos_final' in locals() else pd.DataFrame(),
    'TRATA O TRAFIC PERSONAS': df_trata_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_trata_final' in locals() else pd.DataFrame(),
    'CODIGO OPERATIVO': df_codigo_operativo_final.reset_index(drop=True).fillna("").replace("", "-") if 'df_codigo_operativo_final' in locals() else pd.DataFrame(),
}

# Cargar la plantilla
wb = load_workbook(TEMPLATE_PATH)

# Escribir los datos
for sheet_name, df in dataframes.items():
    sheet = wb[sheet_name]
    for i, row in df.iterrows():
        for j, value in enumerate(row):
            cell= sheet.cell(row=i+4, column=j+2, value=value) # type: ignore
            cell.alignment = Alignment(horizontal='center')
            

for sheet in wb.worksheets:
    sheet.conditional_formatting = []
    sheet.auto_filter.ref = None
    sheet.auto_filter = None
# Guardar el archivo
output_file = f'informes/Informe_{FECHA_MINIMA.replace("-","_")}_AL_{FECHA_MAYOR.replace("-","_")}.xlsx'

wb.save(output_file)
print(f"Archivo generado: {output_file}")



# ### GUARDAR EN LA BASE DE DATOS LA INFORMACION ACTUALIZADA

# In[25]:


if ACTUALIZAR_BASE:
    try:
        # Intentar leer el archivo base
        if os.path.exists(PATH_BASE):
            wb_base = load_workbook(PATH_BASE)
        else:
            raise FileNotFoundError

    except FileNotFoundError:
        # Crear un archivo nuevo si no existe
        print(f"Archivo '{PATH_BASE}' no encontrado. Creando uno nuevo.")
        wb_base = Workbook()
        wb_base.save(PATH_BASE)
        print(f"Archivo '{PATH_BASE}' creado correctamente.")

    except Exception as e:
        # Manejar otros errores
        print(f"Error al leer el archivo '{PATH_BASE}': {e}")
        wb_base = Workbook()
        wb_base.save(PATH_BASE)

    # Procesar hojas y escribir datos
    for sheet_name, df in dataframes.items():
        if sheet_name in wb_base.sheetnames:
            sheet = wb_base[sheet_name]
            last_row = sheet.max_row
            # Escribir títulos si la hoja está vacía
            if last_row == 0:
                for col_num, column_title in enumerate(df.columns, start=1):
                    cell = sheet.cell(row=1, column=col_num, value=column_title)
                    cell.alignment = Alignment(horizontal='center')
                last_row = 1
        else:
            # Crear hoja nueva
            sheet = wb_base.create_sheet(title=sheet_name)
            for col_num, column_title in enumerate(df.columns, start=1):
                cell = sheet.cell(row=1, column=col_num, value=column_title)
                cell.alignment = Alignment(horizontal='center')
            last_row = 1

        # Escribir los datos debajo de la última fila
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                cell = sheet.cell(row= last_row + i + 1, column=j + 1, value=value) # type: ignore
                cell.alignment = Alignment(horizontal='center')

    # Guardar los cambios
    wb_base.save(PATH_BASE)
    print(f"Archivo base actualizado: {PATH_BASE}")


    
print("\n".join(f"{k}: {v}" for k, v in CONTADOR.items()))

